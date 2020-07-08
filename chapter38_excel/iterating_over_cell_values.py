# iterating_over_cell_values.py

from openpyxl import load_workbook

def iterating_over_values(path):
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    for value in sheet.iter_rows(min_row=1, max_row=3,
                                 min_col=1, max_col=3,
                                 values_only=True):
        print(value)

if __name__ == '__main__':
    iterating_over_values('books.xlsx')