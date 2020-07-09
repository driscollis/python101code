# remove_sheets.py

import openpyxl

def create_worksheets(path):
    workbook = openpyxl.Workbook()
    sheet1 = workbook.create_sheet()
    # Insert a worksheet
    workbook.create_sheet(index=1,
                          title='Second sheet')
    print(workbook.sheetnames)
    workbook.remove(workbook['Second sheet'])
    print(workbook.sheetnames)
    workbook.save(path)

if __name__ == '__main__':
    create_worksheets('remove_sheets.xlsx')