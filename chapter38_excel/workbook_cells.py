# workbook_cells.py

from openpyxl import load_workbook

def get_cell_info(path):
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    print(sheet)
    print(f'The title of the Worksheet is: {sheet.title}')
    print(f'The value of {sheet["A2"].value=}')
    print(f'The value of {sheet["A3"].value=}')
    cell = sheet['B3']
    print(f'{cell.value=}')

def get_info_by_coord(path):
    workbook = load_workbook(filename=path)
    sheet = workbook.active
    cell = sheet['A2']
    print(f'Row {cell.row}, Col {cell.column} = {cell.value}')
    print(f'{cell.value=} is at {cell.coordinate=}')

if __name__ == '__main__':
    get_cell_info('books.xlsx')
    get_info_by_coord('books.xlsx')